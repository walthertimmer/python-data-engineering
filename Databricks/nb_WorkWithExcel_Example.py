# Databricks notebook source
# params
workspace_name = "xxx-123456789101112"

# COMMAND ----------

# MAGIC %md
# MAGIC ## Excel

# COMMAND ----------

workspace = spark.conf.get("spark.databricks.workspaceUrl").split('.')[0]

# COMMAND ----------

try:
    spark.catalog.listCatalogs()
except Exception as e:
    print(e)

# COMMAND ----------

# basic params
if workspace == workspace_name: #01
    catalog_name = "hive_metastore"
    schema_name = "demo"
    volume_name = "demo_volume"
elif workspace == workspace_name: #p1
    catalog_name = "hive_metastore"
    schema_name = "test"
    volume_name = "test_volume"

print(f"{catalog_name} {schema_name} {volume_name}")

# COMMAND ----------

# MAGIC %md ##### Get the basics in place

# COMMAND ----------

try:
    spark.sql(f"USE CATALOG {catalog_name};")
except Exception as e:
    print(e)

# COMMAND ----------

try:
    spark.sql(f"CREATE SCHEMA IF NOT EXISTS {catalog_name}.{schema_name}   ")
except Exception as e:
    print(e)

# COMMAND ----------

try:
  spark.sql(f"CREATE VOLUME IF NOT EXISTS {catalog_name}.{schema_name}.{volume_name}   ")
except Exception as e:
  print(e)
  try:
    spark.sql(f"CREATE VOLUME IF NOT EXISTS {schema_name}.{volume_name}   ")
  except Exception as e:
    print(e)

# COMMAND ----------

# MAGIC %md
# MAGIC ## PANDAS

# COMMAND ----------

# MAGIC %md 
# MAGIC #### Read from excel
# MAGIC

# COMMAND ----------

excel_path = f"/Volumes/{catalog_name}/{schema_name}/{volume_name}/test/Voorbeeld_Excel.xlsx"

#df_excel = spark.read.format("com.crealytics.spark.excel").option("header","true").option("inferSchema","true").load(excel_path)

# COMMAND ----------

import pandas as pd

# first read it with pandas instead of spark native because package is not installed issue.. then convert to spark dataframe for better distribution
pd_excel = pd.read_excel(excel_path)
df_excel = spark.createDataFrame(pd_excel)

# COMMAND ----------

display(df_excel)

# COMMAND ----------

# MAGIC %md
# MAGIC #### Write to excel
# MAGIC Note that writing a file directly to dbfs is not supported since you cannot keep a file open while writing. So you first need to create the file locally on cluster and then copy/write it to DBFS

# COMMAND ----------

# add some data
columns = ['a','b']
values = [
    (101,102),
    (201,202),
]
df_NewData = spark.createDataFrame(values,columns)
display(df_NewData)

# COMMAND ----------

# DBTITLE 1,combine some data
df_excel = df_excel.union(df_NewData)
display(df_excel)

# COMMAND ----------

# MAGIC %md
# MAGIC #### Write away the file

# COMMAND ----------

# write locally
pd_excel = df_excel.toPandas()

# multiple sheets
with pd.ExcelWriter('output_excel_file.xlsx') as writer:
    pd_excel.to_excel(writer, sheet_name='Output', header=True,index=False)
    pd_excel.to_excel(writer, sheet_name='OutputSheet2', header=True,index=False)

# COMMAND ----------

# MAGIC %sh
# MAGIC # list all files in current directory
# MAGIC ls -a
# MAGIC pwd

# COMMAND ----------

import os
os.environ['EXCEL_PATH'] = f'/Volumes/{catalog_name}/{schema_name}/{volume_name}/test/Export_Excel.xlsx'

# COMMAND ----------

# MAGIC %sh
# MAGIC # copy data from local cluster to dbfs
# MAGIC mv output_excel_file.xlsx $EXCEL_PATH

# COMMAND ----------

# MAGIC %md
# MAGIC ## OPENPYXL

# COMMAND ----------

pip install openpyxl

# COMMAND ----------

# MAGIC %md
# MAGIC #### Formules in excel
# MAGIC You also want to be able not just to write hard values but also the formulas itself to excel

# COMMAND ----------

# add a formule in column c
import openpyxl

workbook = openpyxl.load_workbook(excel_path)
# sheet = workbook.get_sheet_by_name('Blad1')
sheet = workbook['Blad1']

sheet['C1'] = 'Databricks'
sheet['C2'] = '=sum(A2:B2)'

workbook.save('excel_met_formule.xlsx')

# COMMAND ----------

# MAGIC %md
# MAGIC add more data using openpyxl

# COMMAND ----------

import random

i = 50
while i > 2:
    sheet[f'A{i}'] = random.randint(1,999999)
    sheet[f'B{i}'] = random.randint(1,999999)
    #print(i)
    i = i - 1

workbook.save('excel_met_formule.xlsx')

# COMMAND ----------

# MAGIC %md
# MAGIC There is also a possibility you want to add formules using loops

# COMMAND ----------

i = 50
while i > 1: # don't overwrite the header row
  sheet[f'C{i}'] = f'=sum(A{i}:B{i})'
  i = i - 1

workbook.save('excel_met_formule.xlsx')

# COMMAND ----------

# MAGIC %md
# MAGIC get it to volume

# COMMAND ----------

import os
os.environ['EXCEL_PATH'] = f'/Volumes/{catalog_name}/{schema_name}/{volume_name}/test/excel_met_formule.xlsx'

# COMMAND ----------

# MAGIC %sh
# MAGIC # copy data from local cluster to dbfs
# MAGIC mv excel_met_formule.xlsx $EXCEL_PATH

# COMMAND ----------

# MAGIC %md
# MAGIC Dus alles wat je in Excel zou willen doen zoals tabblad beveiliging, conditionele opmaak

# COMMAND ----------

# list content of excel
list(sheet.values)

# COMMAND ----------

# MAGIC %md
# MAGIC ### Conditional markup (formatting)
# MAGIC
# MAGIC https://openpyxl.readthedocs.io/en/stable/formatting.html

# COMMAND ----------

from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# COMMAND ----------

workbook = openpyxl.load_workbook(f'/Volumes/{catalog_name}/{schema_name}/{volume_name}/test/excel_met_formule.xlsx')

workbook.create_sheet('Blad2')
#worksheet = workbook['Blad2']

# COMMAND ----------

# Create fill
red_fill = PatternFill(
    start_color='EE1111',
    end_color='EE1111',
    fill_type='solid'
    )


# COMMAND ----------

# Add a conditional formatting based on a cell comparison
# addCellIs(range_string, operator, formula, stopIfTrue, wb, font, border, fill)
# Format if cell is less than 'formula'

workbook['Blad2'].conditional_formatting.add(
    'A1:A10',
    CellIsRule(operator='lessThan', formula=['10'], stopIfTrue=True, fill=red_fill)
    )



# COMMAND ----------

sheet = workbook['Blad2']
sheet['A2'] = '50'
sheet['A5'] = '5'

# COMMAND ----------

workbook.save('excel_met_formule.xlsx')

# COMMAND ----------

import os
os.environ['EXCEL_PATH'] = f'/Volumes/{catalog_name}/{schema_name}/{volume_name}/test/excel_met_formule.xlsx'

# COMMAND ----------

# MAGIC %sh
# MAGIC # copy data from local cluster to dbfs
# MAGIC mv excel_met_formule.xlsx $EXCEL_PATH

# COMMAND ----------

# MAGIC %md
# MAGIC ## Security  
# MAGIC https://openpyxl.readthedocs.io/en/stable/protection.html

# COMMAND ----------

# MAGIC %md
# MAGIC ### Workbook protection

# COMMAND ----------

# To prevent other users from viewing hidden worksheets, adding, moving, deleting, or hiding worksheets, and renaming worksheets, you can protect the structure of your workbook with a password. The password can be set using the openpyxl.workbook.protection.WorkbookProtection.workbookPassword() property

# Similarly removing change tracking and change history from a shared workbook can be prevented by setting another password. This password can be set using the openpyxl.workbook.protection.WorkbookProtection.revisionsPassword() property

# from openpyxl import Workbook
from openpyxl.workbook.protection import WorkbookProtection

workbook = openpyxl.load_workbook(f'/Volumes/{catalog_name}/{schema_name}/{volume_name}/test/excel_met_formule.xlsx')

workbook.security = WorkbookProtection(
    lockStructure = True, 
    lockRevision=True
    )
    #    revisionsPassword = 'walther'
    #      workbookPasswordCharacterSet = 'walther', 


workbook.security.set_workbook_password('walther', already_hashed=False)

workbook.save('excel_met_formule.xlsx')

# COMMAND ----------

# MAGIC %sh
# MAGIC # copy data from local cluster to dbfs
# MAGIC mv excel_met_formule.xlsx $EXCEL_PATH

# COMMAND ----------

# MAGIC %md
# MAGIC ### Worksheet protection

# COMMAND ----------

# Various aspects of a worksheet can also be locked by setting attributes on the openpyxl.worksheet.protection.SheetProtection object. Unlike workbook protection, sheet protection may be enabled with or without using a password. Sheet protection is enabled using the openpxyl.worksheet.protection.SheetProtection.sheet attribute or calling enable() or disable():

from openpyxl.worksheet.protection import SheetProtection

worksheet = workbook['Blad1']
worksheet.security = worksheet.protection.sheet = True
worksheet.security = worksheet.protection.password = 'walther'
worksheet.security = worksheet.protection.enable()
# worksheet.protection.disable()

# COMMAND ----------

workbook.save('excel_met_formule.xlsx')

# COMMAND ----------

# MAGIC %sh
# MAGIC # copy data from local cluster to dbfs
# MAGIC mv excel_met_formule.xlsx $EXCEL_PATH
